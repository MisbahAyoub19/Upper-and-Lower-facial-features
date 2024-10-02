import os
from openpyxl import load_workbook
import csv

# Path to the folder containing extracted features
features_folder = 'D:/Pycharm_result/Emotion/ADFES_result_previous/sheet/'

# Define the columns to extract
columns_to_extract =  ['frameno', 'Sub_P1', 'Sub_P2', 'Sub_P3', 'Sub_P4', 'Sub_P5',
                      'REB_P1', 'REB_P2', 'REB_P3', 'REB_P4', 'REB_P5']
output_file = 'last_10_features_nose.csv'

# Function to process each sheet in the Excel file
def process_sheet(sheet_path):
    # Load the Excel sheet
    wb = load_workbook(sheet_path)
    sheet = wb.active

    # Get the last 10 rows
    last_ten_rows = list(sheet.iter_rows(max_row=sheet.max_row, min_row=max(sheet.max_row - 9, 1),
                                         max_col=sheet.max_column, min_col=1, values_only=True))[-10:]

    # Select only the specified columns
    selected_rows = [[row[i - 1] for i in range(1, len(columns_to_extract) + 1)] for row in last_ten_rows]

    return selected_rows

# Function to process each file in the features folder
def process_files(folder_path):
    # Create an empty list to store the results
    result_rows = []

    # Iterate through each file in the folder
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.xlsx'):
                # Construct the full path to the Excel file
                file_path = os.path.join(root, file)

                # Process the sheet and append the result to the list
                result_rows.extend(process_sheet(file_path))

    return result_rows

# Main function
def main():
    # Process the files in the features folder
    result_rows = process_files(features_folder)

    # Write the result to a CSV file
    with open(output_file, 'w', newline='') as csvfile:
        csv_writer = csv.writer(csvfile)

        # Write the header
        csv_writer.writerow(columns_to_extract)

        # Write the rows
        csv_writer.writerows(result_rows)

if __name__ == "__main__":
    main()
