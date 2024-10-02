import glob
import cv2
import os
#import accuracy
import time
import pandas as pd

import scipy.ndimage as ndi
import pylab as pl
import matplotlib.patches as mpatches
from PIL import Image
import sys
from xlwt import *
import xlwt


from skimage import data
from skimage.filters import threshold_otsu
from skimage.segmentation import clear_border
from skimage.measure import label, regionprops
from skimage.morphology import closing, square
from skimage.color import label2rgb
from skimage import io,measure,color,data,filters
import matplotlib.image as mpimg

from matplotlib import pyplot as plt
from openpyxl import Workbook
import scipy.stats as stats
import pandas as pd
import xlsxwriter
import roi_v2
#import packages for accuracy
from sklearn.ensemble import RandomForestClassifier
from sklearn import metrics
from sklearn.model_selection import train_test_split
from sklearn import datasets, linear_model
from matplotlib import pyplot as plt
import numpy as np
from sklearn.metrics import confusion_matrix
import seaborn as sns
sns.set_style('darkgrid')

#import r_eye
#import r_eyebrow
#import l_eyebrow
start = time.process_time()
from matplotlib.backends.backend_pdf import PdfPages
# workbook = xlsxwriter.Workbook(outFileXLSX)
# worksheet1 = workbook.add_worksheet('Results 1.xlsx')
# worksheet2 = workbook.add_worksheet('Results 2')
# sheet1 = worksheet1.active()
# sheet2 = worksheet2.active()

workbook = Workbook()
workbook1 = Workbook()
sheet = workbook.active
workbook1.create_sheet('sheet1')
sheet1 = workbook1.active






# file = xlwt.Workbook()
# sheet = file.add_sheet('Sheet1', cell_overwrite_ok=True)
#sheet = file.active
#cols = ["A", "B", "C", "D", "E"]
#txt = [0,1,2,3,4]
# column name
sheet1["A1"] = "frame_list"
sheet1["B1"] = "blink_ratio"
sheet1["C1"] = "blink_actual"
sheet1["D1"] = "blink_predicted"
sheet1["E1"] = "partial_actual"
sheet1["F1"] = "partial_predicte"
sheet1["G1"] = "no_blink_actual"
sheet1["H1"] = "no_blink_predicted"
sheet1["I1"] = "accuracy"
sheet1["J1"] = "percision_Positive"
sheet1["K1"] = "percision_negative"
sheet1["L1"] = "recall_sensitivity"
sheet1["M1"] = "recall_specifity"
sheet1["N1"] = "f1_positive"
sheet1["O1"] = "f1_negative"
sheet1["P1"] = "haming loss"
sheet1["Q1"] = "jacard score"
sheet1["R1"] = "cross entropy loss"



sheet["A1"] = "Frame no"
sheet["B1"] = "Blink Ratio"
sheet["C1"] = "Left Eye ratio"
sheet["D1"] = "Right Eye Ratio"
sheet["E1"] = "Sub_P1" #LEB stands for left eyebrow
sheet["F1"] = "Sub_P2"
sheet["G1"] = "Sub_P3"
sheet["H1"] = "Sub_P4"
sheet["I1"] = "Sub_P5"
sheet["J1"] = "REB_P1"
sheet["K1"] = "REB_P2"
sheet["L1"] = "REB_P3"
sheet["M1"] = "REB_P4"
sheet["N1"] = "REB_P5"
sheet["O1"] = "Required Time"

sheet["P1"] = "MAX_Ave_Left"
sheet["Q1"] = "MIN_AVe_left"
sheet["R1"] = "MAX_AVe_Right"
sheet["S1"] = "MIN_AVe_Right"
sheet["T1"] = "N_Blink ratio"
sheet["U1"] = "N_left Blink"
sheet["V1"] = "N_right Blink"
sheet["W1"] = "Mouth_Ratio"
sheet["X1"] = "total counts"
sheet["Y1"] = "Blink Count"
sheet["Z1"] = "Partial Blink"
sheet["AA1"] = "No blink"
sheet["AB1"] = "Processing Time"
sheet["AC1"] = "Total"
sheet["AD1"] = "Blink count"
sheet["AE1"] = "Partial blink"
sheet["AF1"] = "No_Blink"
sheet["AG1"] = "Defective Frames"
sheet["AH1"] = "Total Frames"
sheet["AI1"] = "Normalized Mouth ratio"
sheet["AJ1"] = "Norlamized Nose Ratio"
sheet["AK1"] = "Nose ratio"



# sheet.write('A1', 'Frame No')
# sheet.write('B1', 'Blink Ratio')
# sheet.write('C1', 'Left Eye Ratio')
# sheet.write('D1', 'Right Eye ratio')


def Frame(detector,predictor, shapePoints, borders, Video_path,PicturePath,imgPaths,GaborPath,SheetPath,FeaturesPath,Gabor_L_EBPath,ResultPath,vidFolder, frame_rate):

 # Frame, function that does everything
 #print('in frame')

 if not os.path.exists(PicturePath):
        # creates picture folder
        os.mkdir(os.path.join(PicturePath))
        # Create a folder for saving graphs
 graph_folder = os.path.join(vidFolder, 'graph')
 if not os.path.exists(graph_folder):
     os.mkdir(graph_folder)



 # for each video in folder
 for video in glob.glob(Video_path): # path of videos
    (filepath, tempfilename) = os.path.split(video)
    (shotname, extension) = os.path.splitext(tempfilename)
    folder_name = shotname
    # Creates path for individual video
    Path = os.path.join(PicturePath, folder_name)
    if not os.path.exists(Path):
        os.mkdir(Path)

    if not os.path.exists(SheetPath):
        os.mkdir(os.path.join(SheetPath))

    cur_dir = os.path.join(SheetPath, shotname)
    if not os.path.exists(cur_dir):
        os.mkdir(cur_dir)

    # reads video and stores in cap variable
    cap = cv2.VideoCapture(video)
    fps = cap.get(cv2.CAP_PROP_FPS)
    size = (int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)), int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)))

    #size=(960,544)
    size= (640,480)

    print ('size', size)



    # Create empty list for frames and for values
    # Note, create it OUTSIDE loop, so it doesn't get overwritten
    v_eyeFrame = []
    v_eyeBlink = []
    v_eyeLeft = []
    v_eyeRight = []
    v_mouth_ratio=[]
    v_new_mouth_ratio = [] #mouth ratio distance
    v_nose_ratio = []
    v_new_nose_ratio = []  #nose ratio for distance
    lefteye_point_distance=[] #distance from eyebrow to eyes(have 5 points)
    righteye_point_distance=[] ##distance from eyebrow to eyes(have 5 points)
    #lists fo that points that can store the 5 points values

    frame_list = []  # this is the list of frames created for the program skipped with defected frames(eyebrow movemnt)
    new_blink = []
    new_left_blink = []
    new_right_blink = []

    z_new_blink= []   # normalized value list
    z_left_blink = []
    z_right_blink = []



    pl1 = []   #points for left eye
    pl2 = []
    pl3 = []
    pl4 = []
    pl5 = []
    pr1 = [] #points for right eye
    pr2 = []
    pr3 = []
    pr4 = []
    pr5 = []

    sub_pl1 =[] # subtraction values fpr left eyesl
    sub_pl2 = []
    sub_pl3 = []
    sub_pl4 = []
    sub_pl5= []
    sub_lpoints=[sub_pl1,sub_pl2, sub_pl3,sub_pl4,sub_pl5]


    point_lists = [pl1, pl2, pl3, pl4, pl5]
    point_lists2 = [pr1, pr2, pr3, pr4, pr5]
    errorlist =[4.5,4.5,4.5,4.5,4.5]



    # for each frame in cap, read
    i = 0
    while (cap.isOpened()):  # cv2.VideoCapture.isOpened()
        i = i + 1
        ret, frame = cap.read()  # cv2.VideoCapture.read()　
        if ret == True:
            path = Path +'/'
            picturepath = path+ str('%02d' % i) + '.jpg'

            #print (picturepath)
            # Apply horizontal flip augmentation
            frame = cv2.flip(frame, 1)

            # Writes all raw pictures to file
            cv2.imwrite(picturepath, frame)

            # Save duplication by just calling detector once
            dets = detector(frame, 1)
            lenDets = len(dets)

            if(lenDets>0):
                for featCount in range(len(shapePoints)):
                    #print ("frame is ", i)

                    #print(len(dets))

                # Extracts right eye and gabor of feature
                # do this for each frame
                #print('inside a frame')
                # Modify the call to return 2 values
                    (framOut, blinkOut, left_eyeratio, right_eyeratio, left_eb_distance, right_eb_distance, mouth_distance, nose_distance) = roi_v2.rect1(predictor, i, shotname, picturepath,
                                                                   imgPaths[featCount], GaborPath, SheetPath, FeaturesPath, shapePoints[featCount], borders[featCount], frame,
                                                                   dets)

                #append values to list
                leftlist = list(left_eb_distance)
                rightlist = list(right_eb_distance)
                v_eyeFrame.append(framOut)
                v_eyeBlink.append(blinkOut)
                v_eyeLeft.append(left_eyeratio)
                v_eyeRight.append(right_eyeratio)
                lefteye_point_distance.append(leftlist)
                righteye_point_distance.append(rightlist)
                v_mouth_ratio.append(mouth_distance)
                v_new_mouth_ratio.append(v_mouth_ratio[0])
                v_nose_ratio.append(nose_distance)
                v_new_nose_ratio.append(v_nose_ratio[0])
                # print(v_new_mouth_ratio)
                # print(v_mouth_ratio, "mouth aspect ratio")
                #print(v_new_nose_ratio, "nose ratio")
                #print(v_new_mouth_ratio, "mouth ratio")


                for z in range(len(v_eyeFrame)):


                    sheet['A' + str(z + 2)].value = v_eyeFrame[z] # Frame values
                    sheet['B' + str(z + 2)].value = v_eyeBlink[z] # eye blink ratio
                    sheet['C' + str(z + 2)].value = v_eyeLeft[z]  # left eye ratio
                    sheet['D' + str(z + 2)].value = v_eyeRight[z]
                    #sheet['J' + str(z + 2)].value = lefteye_point_distance[z]
                    #sheet['K' + str(z + 2)].value = righteye_point_distance[z] # right eye ratio



            else:
                v_eyeFrame.append(len(v_eyeFrame)+1)
                #v_eyeFrame.append(v_eyeFrame)
                v_eyeBlink.append(5.5) # putting these value to identifies defected frames
                v_eyeLeft.append(5.5) # putting 3.5 to idetifies defected frames
                v_eyeRight.append(5.5) #putting these values to identifies defected frames
                lefteye_point_distance.append(errorlist) # putting error list incese of identifying defetced points
                righteye_point_distance.append(errorlist)
                #continue

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break


        else:
            break



    #file.save(cur_dir + '/' + str('%02d') + '.xls')

    #workbook.save(cur_dir + '/'+ shotname +".xlsx")

    ####++++++ for eyebrows frame+++++++

    for i_blink, i_left_blink, i_right_blink in zip(range(len(v_eyeBlink)), range(len(v_eyeLeft)),
                                                    range(len(v_eyeRight))):
        val_blink = v_eyeBlink[i_blink]
        val_left_blink = v_eyeLeft[i_left_blink]
        val_right_blink = v_eyeRight[i_right_blink]
        #print(val_blink, "val_blink")
        if val_blink == 5.5:
            continue

        elif val_left_blink == 5.5:
            continue
        elif val_right_blink == 5.5:
            continue

        new_blink.append(val_blink)
        new_left_blink.append(val_left_blink)
        new_right_blink.append(val_right_blink)



    cap.release()




     #now this loop took the value of every points form every frames and distribute
      # it in % list... p1,p2,p3,p4 p5




    for outer in range(0,5):
        #pointer = 0

        for inerloop1 , inerloop2 in zip((range(len(lefteye_point_distance))),range(len(righteye_point_distance))):
            #print("value of ", inerloop, "and", outer)

            #value = lefteye_point_distance[inerloop1][outer]
            #print(lefteye_point_distance)
            value = lefteye_point_distance[inerloop1][outer]
            value2 = righteye_point_distance[inerloop2][outer]
            if value == 4.5:    # this states would create the problem in eye bliking because we are dealing with eyeblinking defetec frames here
                continue

            elif value2 == 4.5:
                continue


            point_lists[outer].append(value)
            point_lists2[outer].append(value2)






        #This loop would be used if we want to subtracted every points from another means(a2-a1 formula)
    # for index in range(len(point_lists)):
    #     result2 = [x - y for x, y in zip(point_lists2[index], point_lists2[index][1:])]
    #     sub_lpoints[index].append(result)
    #     point_lists2[index].append(result2)
        #print(sub_lpoints[index], "result of ", point_lists[index])



    #include zscore in Numpy array
    z_pl1 = stats.zscore(pl1)  # z denotes z score
    z_pl2 = stats.zscore(pl2)
    z_pl3 = stats.zscore(pl3)
    z_pl4 = stats.zscore(pl4)
    z_pl5 = stats.zscore(pl5)
    z_pr1 = stats.zscore(pr1)
    z_pr2 = stats.zscore(pr2)
    z_pr3 = stats.zscore(pr3)
    z_pr4 = stats.zscore(pr4)
    z_pr5 = stats.zscore(pr5)
    z_v_mouth_ratio = stats.zscore(v_mouth_ratio)
    z_new_blink = stats.zscore(new_blink)
    z_left_blink = stats.zscore(new_left_blink)
    z_right_blink = stats.zscore(new_right_blink)
    z_v_nose_ratio = stats.zscore(v_nose_ratio)

    #print(z_new_blink)

    max_av_left = (max(z_pl1)+max(z_pl2)+max(z_pl3)+max(z_pl4)+max(z_pl5))/5
    min_av_left = ((min(z_pl1)+min(z_pl2)+min(z_pl3)+min(z_pl4)+min(z_pl5))/5)

    max_av_right = (max(z_pr1) + max(z_pr2)+ max(z_pr3)+ max(z_pr4) + max(z_pr5)) / 5
    min_av_right = (min(z_pr1) + min(z_pr2)+ min(z_pr3)+ min(z_pr4)+ min(z_pr5)) / 5

    blink_count_time = time.process_time() - start
    legth = len(v_eyeFrame)
    #print(legth)
    blink = []
    partial = []
    zero_blink=[]

    blink_count = 0
    total = 0
    partial_count = 0
    no_blink = 0
    counter_blink = 0
    counter_p_blink = 0
    counter_no_blink = 0


    for l, nor_blink in enumerate(z_new_blink):  # this is the loop that will run till
        frame_list.append(l)
        #print(new_blink,  "frame list")

        BLINK_RATIO_THRESHOLD = -2.4
        partial_value = -1.4
        if nor_blink > BLINK_RATIO_THRESHOLD and nor_blink <= partial_value:
            partial.append(1)
            blink.append(0)
            zero_blink.append(0)
            partial_count += 1
            total += 1
            counter_blink = 0
            counter_p_blink = 1
            counter_no_blink = 0
            #print(nor_blink, "Partial blink")

        elif nor_blink <= BLINK_RATIO_THRESHOLD:
            #print(nor_blink, "Blink counts here")
            partial.append(0)
            blink.append(1)
            zero_blink.append(0)
            blink_count += 1
            total += 1
            no_blink = 0
            counter_blink = 1
            counter_p_blink = 0
            counter_no_blink = 0

        elif nor_blink > partial_value:


            partial.append(0)
            blink.append(0)
            zero_blink.append(1)
            no_blink += 1
            counter_blink = 0
            counter_p_blink = 0
            counter_no_blink = 1
            total += 1

        total_frame = len(v_eyeFrame)
        defective_frame = len(v_eyeFrame) - total
        sheet['E' + str(l + 2)].value = z_pl1[l]
        sheet["F" + str(l + 2)].value = z_pl2[l]
        sheet["G" + str(l + 2)].value = z_pl3[l]
        sheet["H" + str(l + 2)].value = z_pl4[l]
        sheet["I" + str(l + 2)].value = z_pl5[l]
        sheet["J" + str(l + 2)].value = z_pr1[l]
        sheet["K" + str(l + 2)].value = z_pr2[l]
        sheet["L" + str(l + 2)].value = z_pr3[l]
        sheet["M" + str(l + 2)].value = z_pr4[l]
        sheet["N" + str(l + 2)].value = z_pr5[l]
        sheet["O2"].value = blink_count_time
        sheet["P2"].value = max_av_left
        sheet["Q2"].value = min_av_left
        sheet["R2"].value = max_av_right
        sheet["S2"].value = min_av_right
        sheet["T" + str(l + 2)].value = z_new_blink[l]
        sheet["U" + str(l + 2)].value = z_left_blink[l]
        sheet["V" + str(l + 2)].value = z_right_blink[l]
        sheet["W" + str(l + 2)].value = v_mouth_ratio[l]
        sheet["AI" + str(l + 2)].value = z_v_mouth_ratio[l]
        sheet["AJ" + str(l + 2)].value = z_v_nose_ratio[l]
        sheet["Ak" + str(l + 2)].value = v_nose_ratio[l]
        sheet["X2"].value = total
        sheet["Y" + str(l + 2)].value = blink[l]
        sheet["Z" + str(l + 2)].value = partial[l]
        sheet["AA" + str(l + 2)].value = zero_blink[l]
        sheet1['A' + str(l + 2)].value = frame_list[l]
        sheet1['A' + str(l + 2)].value = frame_list[l] #frame list
        sheet1['B' + str(l + 2)].value = z_new_blink[l] #blink ratio
        sheet1['C' + str(l + 2)].value = blink[l] #actual blink
        sheet1['E' + str(l + 2)].value = partial[l] #actual partial
        sheet1['G' + str(l + 2)].value = zero_blink[l] # actual no blink

        #predicted blink




    prac_no_blink= total-(blink_count+partial_count)
    #defective_frame = len(v_eyeFrame) - total
    # print(total, "total")
    # print(blink_count, "count")
    # print(partial_count, "partial_count")
    # print(no_blink, "no_blink")
    # print(prac_no_blink, "no_blink")

    #print(blink)
    #print(partial)
    #print(zero_blink)
    print(len(v_eyeFrame))
    sheet["AB2"].value = blink_count_time
    sheet["AC2"].value = total
    sheet["AD2"].value = blink_count
    sheet["AE2"].value = partial_count
    sheet["AF2"].value = prac_no_blink
    sheet["AG2"].value = defective_frame
    sheet["AH2"].value = total_frame
    main_directory = 'D:/Pycharm_result/Emotion/5JAN/blink'
    workbook.save(cur_dir + '/' + shotname + ".xlsx")
    #workbook1.save(cur_dir + '/' + shotname +  "accuracy" + ".xlsx")
    #workbook1.save(main_directory + '/' + shotname + "accuracy" + ".xlsx")
    workbook1.save(main_directory + '/' + shotname + "accuracy" + ".xlsx")
    #print(cur_dir, "current direcctory")


    #if we want to add just vlues without normalization then write only pl1, pl2 rather then Z_Pl1 it will store actual value
    # for p in range(0, len(pl1)):
    #     sheet['E' + str(p + 2)].value = z_pl1[p]
    #     sheet["F" + str(p + 2)].value = z_pl2[p]
    #     sheet["G" + str(p + 2)].value = z_pl3[p]
    #     sheet["H" + str(p + 2)].value = z_pl4[p]
    #     sheet["I" + str(p + 2)].value = z_pl5[p]
    #     sheet["J" + str(p + 2)].value = z_pr1[p]
    #     sheet["K" + str(p + 2)].value = z_pr2[p]
    #     sheet["L" + str(p + 2)].value = z_pr3[p]
    #     sheet["M" + str(p + 2)].value = z_pr4[p]
    #     sheet["N" + str(p + 2)].value = z_pr5[p]
    #     sheet["O" + str(p + 2)].value = blink_count
    #     sheet["P" + str(p + 2)].value = max_av_left
    #     sheet["Q" + str(p + 2)].value = min_av_left
    #     sheet["R" + str(p + 2)].value = max_av_right
    #     sheet["S" + str(p + 2)].value = min_av_right
    #     sheet["T" + str(p + 2)].value = z_new_blink[p]
    #     sheet["U" + str(p + 2)].value = z_left_blink[p]
    #     sheet["V" + str(p + 2)].value = z_right_blink[p]
    #     sheet["W" + str(p + 2)].value = v_mouth_ratio[p]
    #
    #
    #
    #
    #     p +=1




        #print(p)

    # #while q < len(result2):
    # for q in range(len(v_eyeFrame)):
    #     sheet["J" + str(q + 2)].value = pr1[q]
    #     sheet["K" + str(q + 2)].value = pr2[q]
    #     sheet["L" + str(q + 2)].value = pr3[q]
    #     sheet["M" + str(q + 2)].value = pr4[q]
    #     sheet["N" + str(q + 2)].value = pr5[q]
    #
    #     q +=1

    #sheet["O"].value = blink_count
    #sheet.write(0,14,blink_count)
    #workbook.write(0,14,blink_count)
    #workbook.save(cur_dir + '/' + shotname + ".xlsx")

    # ####++++++ for eyebrows frame+++++++
    #
    # frame_list = []   # this is the list of frames created for the program skipped with defected frames(eyebrow movemnt)
    # new_blink = []
    # new_left_blink=[]
    # new_right_blink = []
    #
    # #we have creted the loop to store te blinking values with out defected frames so that we can show them in graph
    # for i_blink,i_left_blink,i_right_blink in zip(range(len(v_eyeBlink)),range(len(v_eyeLeft)),range(len(v_eyeRight))):
    #     val_blink = v_eyeBlink[i_blink]
    #     val_left_blink = v_eyeLeft[i_left_blink]
    #     val_right_blink = v_eyeRight[i_right_blink]
    #     if val_blink == 5.5:
    #         continue
    #     elif val_left_blink == 5.5:
    #         continue
    #     elif val_right_blink == 5.5:
    #         continue
    #
    #
    #
    #     new_blink.append(val_blink)
    #     new_left_blink.append(val_left_blink)
    #     new_right_blink.append(val_right_blink)





    #print(new_blink, "new blink")
    #print(new_left_blink, "left blink")
    #print(new_right_blink, "right blink")



    #print(frame_list)
    #print(new_blink)
    #print(new_left_blink)
    #print(new_right_blink)

    # # Save the eye blinking graph
    # plt.figure()
    # plt.plot(v_eyeFrame, z_new_blink, lw=2)
    # plt.title("Eye Blinking")
    # plt.xlabel('Frames')
    # plt.ylabel('Normalized Blink Ratio')
    # plt.savefig(os.path.join(graph_folder, 'f{shotname}_graph_eye_blinking_graph.png'))
    # plt.close()
    #
    # # Save the eyebrow movement graph
    # plt.figure()
    # plt.plot(frame_list, z_pl1, label="point1")
    # plt.plot(frame_list, z_pl2, label="point2")
    # plt.plot(frame_list, z_pl3, label="Point3")
    # plt.plot(frame_list, z_pl4, label="Point4")
    # plt.plot(frame_list, z_pl5, label="Point5")
    # plt.xlabel('Frames')
    # plt.ylabel('Normalized value of Eye Movement')
    # plt.title('Left Eyebrow Movement for normalized value')
    # plt.legend()
    # plt.savefig(os.path.join(graph_folder, 'left_eyebrow_movement_graph.png'))
    # plt.close()
    #
    # # Save the mouth aspect ratio graph
    # plt.figure()
    # plt.plot(v_mouth_ratio, lw=2)
    # plt.title("Mouth Aspect Ratio")
    # plt.xlabel('Frames')
    # plt.ylabel('Mouth Aspect Ratio')
    # plt.savefig(os.path.join(graph_folder, 'mouth_aspect_ratio_graph.png'))
    # plt.close()



 fig = plt.figure(figsize=(9,6))

#***** Gaph for eye blinking
 sub1 = fig.add_subplot(2,3,1) #two rows, two columns, First cell
 sub1.plot(v_eyeFrame,v_eyeBlink)
 sub1.title.set_text('B.R with Defective Frames')
 sub1.set_ylabel('Blink Ratio')
 #sub1.set_xlabel('frames')
 #sub1.legend("res1", loc= 'upper left')


 sub2 = fig.add_subplot(2,3,2) #two rows, two columns, second cell
 sub2.plot(v_eyeFrame,v_eyeLeft, label = "Left Eye ratio")
 sub2.plot(v_eyeFrame,v_eyeRight, label = "right Eye ratio")
 sub2.title.set_text('L&R B.R with D.F') # D.F stands fro defected frames
 sub2.legend()

 sub4 = fig.add_subplot(2,3,3) #two rows, two columns, third cell
 sub4.plot(frame_list,z_left_blink, label = "Left Eye ratio") # new_left_blink replacing  z_left_blink
 sub4.plot(frame_list,z_right_blink, label = "right Eye ratio") # new_right_blink replacing z_left_blink
 sub4.title.set_text('L&R B.R without D.F')
 sub4.legend()

 sub3 = fig.add_subplot(2,2,(3,4)) # two rows, two colums, combined third and fourth cell
 sub3.plot(frame_list,z_new_blink)
 sub3.title.set_text('Eye blinking without Defective Frames') #E.B stands for eye blinking
 sub3.legend()
 plt.xlabel('Frames')
 plt.ylabel('Normalized LAR')
 plt.grid(True)
 # Save the plot
 #fig.savefig(os.path.join(graph_folder, "plot_with_four_subplots.png"))

 graph_file_name_eye = f'{shotname}_eye_blinking_graph.png'
 graph_path_eye = os.path.join(graph_folder, graph_file_name_eye)
 plt.savefig(graph_path_eye)
 plt.close(fig)
 #plt.show()

 # ***** Graph for left and right eyebrow movement
 fig_eyebrows, (ax_left, ax_right) = plt.subplots(2, 1, figsize=(9, 8))

 # Left Eyebrow Movement
 ax_left.plot(frame_list, z_pl1, label="point1")
 ax_left.plot(frame_list, z_pl2, label="point2")
 ax_left.plot(frame_list, z_pl3, label="Point3")
 ax_left.plot(frame_list, z_pl4, label="Point4")
 ax_left.plot(frame_list, z_pl5, label="Point5")
 ax_left.set_title('Left Eyebrow Movement for normalized value')
 ax_left.legend()

 # Right Eyebrow Movement
 ax_right.plot(frame_list, z_pr1, label="point1")
 ax_right.plot(frame_list, z_pr2, label="point2")
 ax_right.plot(frame_list, z_pr3, label="Point3")
 ax_right.plot(frame_list, z_pr4, label="Point4")
 ax_right.plot(frame_list, z_pr5, label="Point5")
 ax_right.set_xlabel('Frames')
 ax_right.set_ylabel('Normalized value of eye Movement ')
 ax_right.set_title('Right Eyebrow movement for normalized value')
 ax_right.legend()

 # Save the eyebrows plot
 graph_file_name_eyebrows = f'{shotname}_eyebrows.png'
 graph_path_eyebrows = os.path.join(graph_folder, graph_file_name_eyebrows)
 fig_eyebrows.savefig(graph_path_eyebrows)
 plt.close(fig_eyebrows)

 # Display the plots
 #plt.show()

 # ***** Graph for mouth aspect ratio
 fig_mouth, ax_mouth = plt.subplots()
 ax_mouth.plot(v_mouth_ratio, lw=2)
 ax_mouth.set_title("Mouth aspect ratio")
 ax_mouth.set_xlabel('Frames')
 ax_mouth.set_ylabel('Mouth aspect ratio')

 # Save the mouth aspect ratio plot
 graph_file_name_mouth = f'{shotname}_mouth.png'
 graph_path_mouth = os.path.join(graph_folder, graph_file_name_mouth)
 fig_mouth.savefig(graph_path_mouth)
 plt.close(fig_mouth)



 # ***** Graph for Nose aspect ratio
 fig_nose, ax_nose = plt.subplots()
 ax_nose.plot(v_nose_ratio, lw=2)
 ax_nose.set_title("Nose aspect ratio")
 ax_nose.set_xlabel('Frames')
 ax_nose.set_ylabel('Nose aspect ratio')

 # Save the mouth aspect ratio plot
 graph_file_name_nose = f'{shotname}_nose.png'
 graph_path_nose = os.path.join(graph_folder, graph_file_name_nose)
 fig_nose.savefig(graph_path_nose)
 plt.close(fig_nose)


 # Display the plots
 #plt.show()

 total_time2 = time.process_time() - start

 print(total_time2, "seconds")
 #plt.show()
 #print(v_mouth_ratio, "vmouth ratio")
 #print(v_new_mouth_ratio, "v new mouth ratio")




 #plt.ioff()
 #accuracy.Frame(SheetPath,shotname,cur_dir,frame_list,blink,partial,zero_blink)
